import json
import os
from collections import defaultdict, Counter
import openpyxl as oxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pgeocode

US_STATE_NAMES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "DC": "District of Columbia", "FL": "Florida", "GA": "Georgia", "HI": "Hawaii",
    "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine",
    "MD": "Maryland", "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota",
    "MS": "Mississippi", "MO": "Missouri", "MT": "Montana", "NE": "Nebraska",
    "NV": "Nevada", "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico",
    "NY": "New York", "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio",
    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "PR": "Puerto Rico",
    "RI": "Rhode Island", "SC": "South Carolina", "SD": "South Dakota",
    "TN": "Tennessee", "TX": "Texas", "UT": "Utah", "VT": "Vermont",
    "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming",
}

CUSTOMERS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "customers.json")
DISTRIBUTORS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "distributors.json")
INSTALLER_EXCEL = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "attached_assets", "Installer_Accounts_Not_On_C4C_1772753485907.xlsx"
)
C4C_EXCEL = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "attached_assets", "Royal_Purple_C4C_Installer_List_1772754933619.xlsx"
)
PROMO_EXCEL = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "attached_assets", "Installer_Promotion_Participation_12.13.25_1772754933620.xlsx"
)

PURPLE = "1a1a2e"
PURPLE_MED = "e31837"
GOLD = "e31837"

HEADER_FILL = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color=PURPLE_MED, end_color=PURPLE_MED, fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color=PURPLE)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color=PURPLE_MED)
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color=PURPLE_MED)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color=PURPLE))

AMBER_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FDF6E3", end_color="FDF6E3", fill_type="solid")
PURPLE_LIGHT_FILL = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
ROSE_FILL = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
INDIGO_FILL = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
EMERALD_FILL = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

TYPE_FILLS = {
    "Promo Only (Not on C4C)": RED_FILL,
    "On Both Lists": GREEN_FILL,
    "C4C Only": BLUE_FILL,
    "Rack Installer": PURPLE_LIGHT_FILL,
    "Distributor": GOLD_FILL,
    "Powersports/Motorsports": ROSE_FILL,
    "International": INDIGO_FILL,
    "Canada": EMERALD_FILL,
}

ALL_TYPES_ORDERED = [
    "Promo Only (Not on C4C)", "On Both Lists", "C4C Only", "Rack Installer",
    "Distributor", "Powersports/Motorsports", "International", "Canada",
]


def _apply_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 30


def _apply_data_row(ws, row, num_cols, alt=False):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        if alt:
            cell.fill = LIGHT_GRAY


def _auto_width(ws, num_cols, max_width=42):
    for col in range(1, num_cols + 1):
        max_len = 10
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 3, max_width)


def _add_auto_filter(ws, header_row, num_cols):
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"
    ws.freeze_panes = f"A{header_row + 1}"


def _type_fill(account_type):
    return TYPE_FILLS.get(account_type, None)


def _write_title(ws, title, subtitle=None, num_cols=8):
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32
    if subtitle:
        ws.merge_cells(f"A2:{get_column_letter(num_cols)}2")
        ws["A2"] = subtitle
        ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="64748B")


def _write_account_table(ws, accounts, start_row, headers, include_type=True, include_rack=False):
    for ci, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=ci, value=h)
    _apply_header_row(ws, start_row, len(headers))
    row = start_row + 1
    for c in accounts:
        col = 1
        ws.cell(row=row, column=col, value=c.get("store_name", c.get("name", ""))); col += 1
        ws.cell(row=row, column=col, value=c.get("address", "")); col += 1
        ws.cell(row=row, column=col, value=c.get("city", "")); col += 1
        ws.cell(row=row, column=col, value=c.get("state", "")); col += 1
        ws.cell(row=row, column=col, value=c.get("county", "")); col += 1
        ws.cell(row=row, column=col, value=c.get("zip", "")); col += 1
        if include_type:
            ws.cell(row=row, column=col, value=c.get("type", "")); col += 1
        if include_rack:
            ws.cell(row=row, column=col, value="Yes" if c.get("rack_installer") else ""); col += 1
        ws.cell(row=row, column=col, value=c.get("country", c.get("type", "US") if "International" in c.get("type", "") else "US")); col += 1
        ws.cell(row=row, column=col, value=c.get("latitude", "")); col += 1
        ws.cell(row=row, column=col, value=c.get("longitude", "")); col += 1

        _apply_data_row(ws, row, len(headers), alt=(row % 2 == 0))
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

        tfill = _type_fill(c.get("type", ""))
        if tfill and include_type:
            type_col = 7
            ws.cell(row=row, column=type_col).fill = tfill
        row += 1

    _auto_width(ws, len(headers))
    _add_auto_filter(ws, start_row, len(headers))
    return row


def _cross_analyze_lists():
    result = {
        "c4c_rows": 0, "c4c_unique": 0, "c4c_dupes": [],
        "promo_rows": 0, "promo_unique": 0, "promo_dupes": [],
        "matched": 0, "only_c4c": 0, "only_promo": 0,
        "matched_names": [], "only_c4c_names": [], "only_promo_names": [],
    }

    c4c_accounts = []
    promo_accounts = []

    if os.path.exists(C4C_EXCEL):
        wb = oxl.load_workbook(C4C_EXCEL)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[1]).strip() if row[1] else ""
            if not name:
                continue
            c4c_accounts.append({
                "name": name,
                "acct_id": str(row[2]).strip() if row[2] else "",
                "street": str(row[3]).strip() if row[3] else "",
                "city": str(row[4]).strip() if row[4] else "",
                "state": str(row[5]).strip() if row[5] else "",
                "zip": str(row[6]).strip() if row[6] else "",
            })

    if os.path.exists(PROMO_EXCEL):
        wb = oxl.load_workbook(PROMO_EXCEL)
        ws = wb['Summary'] if 'Summary' in wb.sheetnames else wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[0]).strip() if row[0] else ""
            if not name:
                continue
            promo_accounts.append({
                "name": name,
                "address": str(row[1]).strip() if row[1] else "",
                "city": str(row[3]).strip() if row[3] else "",
                "state": str(row[4]).strip() if row[4] else "",
                "zip": str(row[5]).strip() if row[5] else "",
            })

    result["c4c_rows"] = len(c4c_accounts)
    result["promo_rows"] = len(promo_accounts)

    c4c_names = set(a["name"].upper() for a in c4c_accounts)
    promo_names = set(a["name"].upper() for a in promo_accounts)

    result["c4c_unique"] = len(c4c_names)
    result["promo_unique"] = len(promo_names)

    matched = c4c_names & promo_names
    only_c4c = c4c_names - promo_names
    only_promo = promo_names - c4c_names

    result["matched"] = len(matched)
    result["only_c4c"] = len(only_c4c)
    result["only_promo"] = len(only_promo)

    c4c_name_counts = Counter(a["name"].upper() for a in c4c_accounts)
    for name, cnt in sorted(c4c_name_counts.items(), key=lambda x: (-x[1], x[0])):
        if cnt <= 1:
            continue
        entries = [a for a in c4c_accounts if a["name"].upper() == name]
        locations = "; ".join(f"{e['city']}, {e['state']}" for e in entries)
        result["c4c_dupes"].append({"name": entries[0]["name"], "count": cnt, "locations": locations})

    promo_name_counts = Counter(a["name"].upper() for a in promo_accounts)
    for name, cnt in sorted(promo_name_counts.items(), key=lambda x: (-x[1], x[0])):
        if cnt <= 1:
            continue
        entries = [a for a in promo_accounts if a["name"].upper() == name]
        locations = "; ".join(f"{e['city']}, {e['state']}" for e in entries[:10])
        if len(entries) > 10:
            locations += f" (+{len(entries)-10} more)"
        result["promo_dupes"].append({"name": entries[0]["name"], "count": cnt, "locations": locations})

    for nm in sorted(matched):
        c4c_entry = next((a for a in c4c_accounts if a["name"].upper() == nm), {})
        promo_entry = next((a for a in promo_accounts if a["name"].upper() == nm), {})
        result["matched_names"].append({
            "name": c4c_entry.get("name", nm),
            "c4c_city": f"{c4c_entry.get('city', '')}, {c4c_entry.get('state', '')}",
            "promo_city": f"{promo_entry.get('city', '')}, {promo_entry.get('state', '')}",
        })

    for nm in sorted(only_c4c):
        entry = next((a for a in c4c_accounts if a["name"].upper() == nm), {})
        result["only_c4c_names"].append({
            "name": entry.get("name", nm),
            "city": entry.get("city", ""),
            "state": entry.get("state", ""),
            "acct_id": entry.get("acct_id", ""),
        })

    for nm in sorted(only_promo):
        entry = next((a for a in promo_accounts if a["name"].upper() == nm), {})
        result["only_promo_names"].append({
            "name": entry.get("name", nm),
            "city": entry.get("city", ""),
            "state": entry.get("state", ""),
        })

    return result


def _get_failed_geolocations():
    if not os.path.exists(INSTALLER_EXCEL):
        return []

    nomi = pgeocode.Nominatim('us')
    wb = oxl.load_workbook(INSTALLER_EXCEL)
    failed = []
    seen = set()

    for sheet_name in ['Not on C4C List', 'Matched Accounts']:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        source = "Not on C4C" if "Not" in sheet_name else "C4C Matched"

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            address = row[1]
            city = row[3]
            state = row[4]
            zipcode = row[5]
            phone = row[6] if len(row) > 6 else None
            email = row[7] if len(row) > 7 else None

            if not name:
                continue
            name = str(name).strip()
            if name in ['-', '(RESIDENCE)'] or name.startswith('('):
                continue

            address = str(address).strip() if address else ""
            city = str(city).strip() if city else ""
            state = str(state).strip().upper() if state else ""
            z = str(zipcode).strip() if zipcode else ""
            z_clean = z.split('-')[0][:5]
            phone = str(phone).strip() if phone else ""
            email = str(email).strip() if email else ""
            if phone == "None":
                phone = ""
            if email == "None":
                email = ""

            key = (name.upper(), city.upper(), state)
            if key in seen:
                continue
            seen.add(key)

            reason = ""
            if not city or not state:
                reason = "Missing city/state"
            elif len(state) != 2 or not state.isalpha():
                reason = f"Invalid state code: {state}"
            else:
                if len(z_clean) == 5 and z_clean.isdigit():
                    result = nomi.query_postal_code(z_clean)
                    if result is None or (result.latitude != result.latitude):
                        reason = f"Zip code not found: {z_clean}"
                    else:
                        continue
                else:
                    reason = f"Invalid zip: {z}"

            failed.append({
                "store_name": name,
                "address": address,
                "city": city,
                "state": state,
                "zip": z,
                "phone": phone,
                "email": email,
                "source": source,
                "reason": reason,
            })

    return failed


def _load_customers():
    if os.path.exists(CUSTOMERS_PATH):
        with open(CUSTOMERS_PATH, "r") as f:
            return json.load(f)
    return []


def _load_distributors():
    if os.path.exists(DISTRIBUTORS_PATH):
        with open(DISTRIBUTORS_PATH, "r") as f:
            data = json.load(f)
            for d in data:
                if "type" not in d:
                    d["type"] = "Distributor"
                if "store_name" not in d and "name" in d:
                    d["store_name"] = d["name"]
            return data
    return []


def generate_c4c_report(output_path):
    customers = _load_customers()
    distributors = _load_distributors()
    all_accounts = customers + distributors

    type_counts = Counter(c.get("type", "Unknown") for c in all_accounts)

    c4c_types = {"Promo Only (Not on C4C)", "On Both Lists", "C4C Only"}
    installer_types = {"Promo Only (Not on C4C)", "On Both Lists", "C4C Only", "Rack Installer"}

    not_c4c = [c for c in all_accounts if c.get("type") == "Promo Only (Not on C4C)"]
    c4c_matched = [c for c in all_accounts if c.get("type") in ("On Both Lists", "C4C Only")]
    rack_installers = [c for c in all_accounts if c.get("type") == "Rack Installer"]
    powersports = [c for c in all_accounts if c.get("type") == "Powersports/Motorsports"]
    international = [c for c in all_accounts if c.get("type") == "International"]
    canada = [c for c in all_accounts if c.get("type") == "Canada"]

    us_accounts = [c for c in all_accounts if c.get("state", "") in US_STATE_NAMES]
    us_states = sorted(set(c["state"] for c in us_accounts))

    county_set = set()
    for c in all_accounts:
        if c.get("county") and c.get("state", "") in US_STATE_NAMES:
            county_set.add((c["state"], c["county"]))

    rack_flagged = sum(1 for c in all_accounts if c.get("rack_installer"))

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════
    # SHEET 1: EXECUTIVE DASHBOARD
    # ═══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = PURPLE

    ws.merge_cells("A1:J1")
    ws["A1"] = "Butler Performance — Complete Account Intelligence Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=PURPLE)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    ws["A2"] = "Prepared by ThrottlePro  |  Complete Network Analysis with C4C Gap Intelligence"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="64748B")

    row = 4
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "What is C4C (Connect for Calumet)?"
    ws[f"A{row}"].font = SUBTITLE_FONT
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = BOTTOM_BORDER
    row += 1

    c4c_explanation = [
        "C4C (Connect for Calumet) is the official dealer/installer onboarding system used by",
        "Butler Performance's parent company, Calumet Specialty Products. When an installer is \"on C4C,\"",
        "they are registered as an authorized Butler Performance dealer — enabling direct pricing,",
        "promotional materials, marketing support, and product updates.",
        "",
        "This report identifies the gap between installer accounts participating in Butler Performance",
        "promotions vs. those registered in C4C, alongside a complete view of all account",
        "categories including distributors, powersports, international, and Canadian partners.",
    ]
    for line in c4c_explanation:
        ws.merge_cells(f"A{row}:J{row}")
        ws[f"A{row}"] = line
        ws[f"A{row}"].font = Font(name="Calibri", size=10, color="475569")
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Network Overview"
    ws[f"A{row}"].font = SUBTITLE_FONT
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = BOTTOM_BORDER
    row += 1

    overview_metrics = [
        ("Total Locations (All Categories)", len(all_accounts), None),
        ("", "", None),
        ("INSTALLER ACCOUNTS", "", None),
        ("  Promo Only (Not on C4C)", type_counts.get("Promo Only (Not on C4C)", 0), RED_FILL),
        ("  On Both Lists (Promo + C4C)", type_counts.get("On Both Lists", 0), GREEN_FILL),
        ("  C4C Only (not on Promo)", type_counts.get("C4C Only", 0), BLUE_FILL),
        ("  Rack Installer (new, unmatched)", type_counts.get("Rack Installer", 0), PURPLE_LIGHT_FILL),
        ("", "", None),
        ("PARTNER NETWORK", "", None),
        ("  Distributors", type_counts.get("Distributor", 0), GOLD_FILL),
        ("  Powersports / Motorsports", type_counts.get("Powersports/Motorsports", 0), ROSE_FILL),
        ("  International", type_counts.get("International", 0), INDIGO_FILL),
        ("  Canada", type_counts.get("Canada", 0), EMERALD_FILL),
        ("", "", None),
        ("COVERAGE", "", None),
        ("  US States (incl. DC/PR)", len(us_states), None),
        ("  US Counties", len(county_set), None),
        ("  Accounts with RP Display Rack", rack_flagged, PURPLE_LIGHT_FILL),
    ]

    for label, val, fill in overview_metrics:
        if not label:
            row += 1
            continue
        if label in ("INSTALLER ACCOUNTS", "PARTNER NETWORK", "COVERAGE"):
            ws.cell(row=row, column=1, value=label).font = SECTION_FONT
            row += 1
            continue
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT if not label.startswith("  ") else DATA_FONT
        cell = ws.cell(row=row, column=2, value=val)
        cell.font = Font(name="Calibri", bold=True, size=11, color=PURPLE)
        cell.alignment = Alignment(horizontal="right")
        if fill:
            ws.cell(row=row, column=1).fill = fill
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "C4C Gap Summary"
    ws[f"A{row}"].font = SUBTITLE_FONT
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = BOTTOM_BORDER
    row += 1

    total_installers = len(not_c4c) + len(c4c_matched)
    gap_pct = len(not_c4c) / total_installers * 100 if total_installers else 0
    matched_pct = len(c4c_matched) / total_installers * 100 if total_installers else 0

    gap_headers = ["Category", "Count", "Percentage"]
    for ci, h in enumerate(gap_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    _apply_header_row(ws, row, 3)
    row += 1

    gap_data = [
        ("Not on C4C (Gap — Needs Onboarding)", len(not_c4c), f"{gap_pct:.1f}%"),
        ("  — On Both Lists (Promo + C4C)", sum(1 for c in c4c_matched if c["type"] == "On Both Lists"),
         f"{sum(1 for c in c4c_matched if c['type'] == 'On Both Lists') / total_installers * 100:.1f}%" if total_installers else "0%"),
        ("  — C4C Only (not on Promo list)", sum(1 for c in c4c_matched if c["type"] == "C4C Only"),
         f"{sum(1 for c in c4c_matched if c['type'] == 'C4C Only') / total_installers * 100:.1f}%" if total_installers else "0%"),
        ("On C4C (Matched)", len(c4c_matched), f"{matched_pct:.1f}%"),
        ("TOTAL Installer Accounts", total_installers, "100%"),
    ]

    for label, cnt, pct in gap_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=cnt)
        ws.cell(row=row, column=3, value=pct)
        _apply_data_row(ws, row, 3, alt=(row % 2 == 0))
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        if label == "TOTAL Installer Accounts":
            for c in range(1, 4):
                ws.cell(row=row, column=c).font = BOLD_FONT
        if label.startswith("  —"):
            ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, color="64748B")
        row += 1

    _auto_width(ws, 3)
    ws.column_dimensions["A"].width = 45

    # ═══════════════════════════════════════════════════════════════
    # SHEET 2: ALL ACCOUNTS (Master List)
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("All Accounts")
    ws2.sheet_properties.tabColor = PURPLE

    _write_title(ws2,
                 f"Complete Account Directory — {len(all_accounts)} Locations",
                 "Every account across all categories. Filter by Type, State, County, or Rack status.",
                 num_cols=11)

    all_sorted = sorted(all_accounts, key=lambda x: (x.get("state", "ZZ"), x.get("county", ""), x.get("store_name", x.get("name", ""))))
    all_headers = ["Store Name", "Address", "City", "State", "County", "Zip",
                   "Account Type", "Rack Installer", "Country", "Latitude", "Longitude"]
    _write_account_table(ws2, all_sorted, 4, all_headers, include_type=True, include_rack=True)

    # ═══════════════════════════════════════════════════════════════
    # SHEET 3: STATE BREAKDOWN (All Types)
    # ═══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("State Breakdown")
    ws3.sheet_properties.tabColor = "e31837"

    _write_title(ws3,
                 "Account Breakdown by State — All Categories",
                 "Every US state with counts per account type, county coverage, and C4C gap percentage.",
                 num_cols=13)

    row = 4
    state_headers = ["State", "Code", "Total", "Counties",
                     "Promo Only", "On Both", "C4C Only", "Rack Inst.",
                     "Distributor", "Powersports", "Gap %", "C4C Rate"]
    for ci, h in enumerate(state_headers, 1):
        ws3.cell(row=row, column=ci, value=h)
    _apply_header_row(ws3, row, len(state_headers))
    row += 1

    state_data = defaultdict(lambda: defaultdict(int))
    state_counties = defaultdict(set)
    for c in all_accounts:
        st = c.get("state", "")
        if st not in US_STATE_NAMES:
            continue
        state_data[st][c.get("type", "Unknown")] += 1
        state_data[st]["_total"] += 1
        if c.get("county"):
            state_counties[st].add(c["county"])

    total_row_data = defaultdict(int)

    for sc in sorted(state_data.keys()):
        d = state_data[sc]
        total = d["_total"]
        promo = d.get("Promo Only (Not on C4C)", 0)
        both = d.get("On Both Lists", 0)
        c4c = d.get("C4C Only", 0)
        rack = d.get("Rack Installer", 0)
        dist = d.get("Distributor", 0)
        ps = d.get("Powersports/Motorsports", 0)
        installer_total = promo + both + c4c
        gap = promo / installer_total * 100 if installer_total else 0
        c4c_rate = (both + c4c) / installer_total * 100 if installer_total else 0
        counties = len(state_counties.get(sc, set()))

        sname = US_STATE_NAMES.get(sc, sc)
        ws3.cell(row=row, column=1, value=sname)
        ws3.cell(row=row, column=2, value=sc)
        ws3.cell(row=row, column=3, value=total)
        ws3.cell(row=row, column=4, value=counties)
        ws3.cell(row=row, column=5, value=promo)
        ws3.cell(row=row, column=6, value=both)
        ws3.cell(row=row, column=7, value=c4c)
        ws3.cell(row=row, column=8, value=rack)
        ws3.cell(row=row, column=9, value=dist)
        ws3.cell(row=row, column=10, value=ps)
        ws3.cell(row=row, column=11, value=f"{gap:.1f}%")
        ws3.cell(row=row, column=12, value=f"{c4c_rate:.1f}%")

        _apply_data_row(ws3, row, len(state_headers), alt=(row % 2 == 0))
        for col in range(2, 13):
            ws3.cell(row=row, column=col).alignment = Alignment(horizontal="center")

        if gap >= 80:
            ws3.cell(row=row, column=11).fill = RED_FILL
            ws3.cell(row=row, column=11).font = Font(name="Calibri", bold=True, size=10, color="DC2626")
        elif gap >= 50:
            ws3.cell(row=row, column=11).fill = AMBER_FILL
        if c4c_rate >= 50:
            ws3.cell(row=row, column=12).fill = GREEN_FILL

        if promo > 50:
            ws3.cell(row=row, column=5).fill = AMBER_FILL

        for k in ["_total", "Promo Only (Not on C4C)", "On Both Lists", "C4C Only",
                   "Rack Installer", "Distributor", "Powersports/Motorsports"]:
            total_row_data[k] += d.get(k, 0)
        total_row_data["counties"] += counties

        row += 1

    t_inst = total_row_data.get("Promo Only (Not on C4C)", 0) + total_row_data.get("On Both Lists", 0) + total_row_data.get("C4C Only", 0)
    t_gap = total_row_data.get("Promo Only (Not on C4C)", 0) / t_inst * 100 if t_inst else 0
    t_rate = (total_row_data.get("On Both Lists", 0) + total_row_data.get("C4C Only", 0)) / t_inst * 100 if t_inst else 0
    ws3.cell(row=row, column=1, value="TOTAL")
    ws3.cell(row=row, column=3, value=total_row_data["_total"])
    ws3.cell(row=row, column=4, value=len(county_set))
    ws3.cell(row=row, column=5, value=total_row_data.get("Promo Only (Not on C4C)", 0))
    ws3.cell(row=row, column=6, value=total_row_data.get("On Both Lists", 0))
    ws3.cell(row=row, column=7, value=total_row_data.get("C4C Only", 0))
    ws3.cell(row=row, column=8, value=total_row_data.get("Rack Installer", 0))
    ws3.cell(row=row, column=9, value=total_row_data.get("Distributor", 0))
    ws3.cell(row=row, column=10, value=total_row_data.get("Powersports/Motorsports", 0))
    ws3.cell(row=row, column=11, value=f"{t_gap:.1f}%")
    ws3.cell(row=row, column=12, value=f"{t_rate:.1f}%")
    for col in range(1, len(state_headers) + 1):
        cell = ws3.cell(row=row, column=col)
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        if col >= 2:
            cell.alignment = Alignment(horizontal="center")

    _auto_width(ws3, len(state_headers))
    _add_auto_filter(ws3, 4, len(state_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 4: COUNTY BREAKDOWN (All Types)
    # ═══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("County Breakdown")
    ws4.sheet_properties.tabColor = "0891B2"

    _write_title(ws4,
                 f"Account Breakdown by County — {len(county_set)} Counties",
                 "Every US county with account counts per type and C4C gap. Filter by State or County to drill down.",
                 num_cols=12)

    row = 4
    county_headers = ["State", "Code", "County", "Total",
                      "Promo Only", "On Both", "C4C Only", "Rack Inst.",
                      "Distributor", "Powersports", "Gap %", "C4C Rate"]
    for ci, h in enumerate(county_headers, 1):
        ws4.cell(row=row, column=ci, value=h)
    _apply_header_row(ws4, row, len(county_headers))
    row += 1

    county_data = defaultdict(lambda: defaultdict(int))
    for c in all_accounts:
        st = c.get("state", "")
        county = c.get("county", "")
        if not st or st not in US_STATE_NAMES or not county:
            continue
        county_data[(st, county)][c.get("type", "Unknown")] += 1
        county_data[(st, county)]["_total"] += 1

    ct_total = defaultdict(int)
    for key in sorted(county_data.keys()):
        st, county = key
        d = county_data[key]
        total = d["_total"]
        promo = d.get("Promo Only (Not on C4C)", 0)
        both = d.get("On Both Lists", 0)
        c4c = d.get("C4C Only", 0)
        rack = d.get("Rack Installer", 0)
        dist = d.get("Distributor", 0)
        ps = d.get("Powersports/Motorsports", 0)
        inst_total = promo + both + c4c
        gap = promo / inst_total * 100 if inst_total else 0
        c4c_rate = (both + c4c) / inst_total * 100 if inst_total else 0

        sname = US_STATE_NAMES.get(st, st)
        ws4.cell(row=row, column=1, value=sname)
        ws4.cell(row=row, column=2, value=st)
        ws4.cell(row=row, column=3, value=county)
        ws4.cell(row=row, column=4, value=total)
        ws4.cell(row=row, column=5, value=promo)
        ws4.cell(row=row, column=6, value=both)
        ws4.cell(row=row, column=7, value=c4c)
        ws4.cell(row=row, column=8, value=rack)
        ws4.cell(row=row, column=9, value=dist)
        ws4.cell(row=row, column=10, value=ps)
        ws4.cell(row=row, column=11, value=f"{gap:.1f}%" if inst_total else "N/A")
        ws4.cell(row=row, column=12, value=f"{c4c_rate:.1f}%" if inst_total else "N/A")

        _apply_data_row(ws4, row, len(county_headers), alt=(row % 2 == 0))
        for col in range(2, 13):
            ws4.cell(row=row, column=col).alignment = Alignment(horizontal="center")

        if inst_total:
            if gap >= 80:
                ws4.cell(row=row, column=11).fill = RED_FILL
            elif gap >= 50:
                ws4.cell(row=row, column=11).fill = AMBER_FILL
            if c4c_rate >= 50:
                ws4.cell(row=row, column=12).fill = GREEN_FILL

        for k in d:
            ct_total[k] += d[k]
        row += 1

    ct_inst = ct_total.get("Promo Only (Not on C4C)", 0) + ct_total.get("On Both Lists", 0) + ct_total.get("C4C Only", 0)
    ct_gap_pct = ct_total.get("Promo Only (Not on C4C)", 0) / ct_inst * 100 if ct_inst else 0
    ct_c4c_pct = (ct_total.get("On Both Lists", 0) + ct_total.get("C4C Only", 0)) / ct_inst * 100 if ct_inst else 0
    ws4.cell(row=row, column=1, value="TOTAL")
    ws4.cell(row=row, column=3, value=f"{len(county_data)} counties")
    ws4.cell(row=row, column=4, value=ct_total["_total"])
    ws4.cell(row=row, column=5, value=ct_total.get("Promo Only (Not on C4C)", 0))
    ws4.cell(row=row, column=6, value=ct_total.get("On Both Lists", 0))
    ws4.cell(row=row, column=7, value=ct_total.get("C4C Only", 0))
    ws4.cell(row=row, column=8, value=ct_total.get("Rack Installer", 0))
    ws4.cell(row=row, column=9, value=ct_total.get("Distributor", 0))
    ws4.cell(row=row, column=10, value=ct_total.get("Powersports/Motorsports", 0))
    ws4.cell(row=row, column=11, value=f"{ct_gap_pct:.1f}%")
    ws4.cell(row=row, column=12, value=f"{ct_c4c_pct:.1f}%")
    for col in range(1, len(county_headers) + 1):
        cell = ws4.cell(row=row, column=col)
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        if col >= 2:
            cell.alignment = Alignment(horizontal="center")

    _auto_width(ws4, len(county_headers))
    _add_auto_filter(ws4, 4, len(county_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 5: NOT ON C4C (Full List)
    # ═══════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Not on C4C")
    ws5.sheet_properties.tabColor = "DC2626"

    _write_title(ws5,
                 f"Installer Accounts NOT on C4C — {len(not_c4c)} Accounts (Gap)",
                 "These accounts participate in RP promotions but have NOT been onboarded to C4C. Priority for C4C registration.",
                 num_cols=10)

    nc_headers = ["Store Name", "Address", "City", "State", "County", "Zip",
                  "Rack Installer", "Country", "Latitude", "Longitude"]
    nc_sorted = sorted(not_c4c, key=lambda x: (x.get("state", ""), x.get("county", ""), x.get("store_name", "")))
    row = 4
    for ci, h in enumerate(nc_headers, 1):
        ws5.cell(row=row, column=ci, value=h)
    _apply_header_row(ws5, row, len(nc_headers))
    row += 1
    for c in nc_sorted:
        ws5.cell(row=row, column=1, value=c.get("store_name", ""))
        ws5.cell(row=row, column=2, value=c.get("address", ""))
        ws5.cell(row=row, column=3, value=c.get("city", ""))
        ws5.cell(row=row, column=4, value=c.get("state", ""))
        ws5.cell(row=row, column=5, value=c.get("county", ""))
        ws5.cell(row=row, column=6, value=c.get("zip", ""))
        ws5.cell(row=row, column=7, value="Yes" if c.get("rack_installer") else "")
        ws5.cell(row=row, column=8, value=c.get("country", "US"))
        ws5.cell(row=row, column=9, value=c.get("latitude", ""))
        ws5.cell(row=row, column=10, value=c.get("longitude", ""))
        _apply_data_row(ws5, row, len(nc_headers), alt=(row % 2 == 0))
        ws5.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        row += 1
    _auto_width(ws5, len(nc_headers))
    _add_auto_filter(ws5, 4, len(nc_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 6: C4C MATCHED (Full List)
    # ═══════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("C4C Matched")
    ws6.sheet_properties.tabColor = "16A34A"

    _write_title(ws6,
                 f"Installer Accounts on C4C — {len(c4c_matched)} Accounts",
                 "Accounts registered in the C4C system. Includes both matched (also on Promo list) and C4C-only accounts.",
                 num_cols=11)

    mc_headers = ["Store Name", "Address", "City", "State", "County", "Zip",
                  "C4C Status", "Rack Installer", "Country", "Latitude", "Longitude"]
    mc_sorted = sorted(c4c_matched, key=lambda x: (x.get("state", ""), x.get("county", ""), x.get("store_name", "")))
    row = 4
    for ci, h in enumerate(mc_headers, 1):
        ws6.cell(row=row, column=ci, value=h)
    _apply_header_row(ws6, row, len(mc_headers))
    row += 1
    for c in mc_sorted:
        ws6.cell(row=row, column=1, value=c.get("store_name", ""))
        ws6.cell(row=row, column=2, value=c.get("address", ""))
        ws6.cell(row=row, column=3, value=c.get("city", ""))
        ws6.cell(row=row, column=4, value=c.get("state", ""))
        ws6.cell(row=row, column=5, value=c.get("county", ""))
        ws6.cell(row=row, column=6, value=c.get("zip", ""))
        ws6.cell(row=row, column=7, value=c.get("type", ""))
        ws6.cell(row=row, column=8, value="Yes" if c.get("rack_installer") else "")
        ws6.cell(row=row, column=9, value=c.get("country", "US"))
        ws6.cell(row=row, column=10, value=c.get("latitude", ""))
        ws6.cell(row=row, column=11, value=c.get("longitude", ""))
        _apply_data_row(ws6, row, len(mc_headers), alt=(row % 2 == 0))
        ws6.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        tfill = _type_fill(c.get("type", ""))
        if tfill:
            ws6.cell(row=row, column=7).fill = tfill
        row += 1
    _auto_width(ws6, len(mc_headers))
    _add_auto_filter(ws6, 4, len(mc_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 7: TOP PRIORITY STATES
    # ═══════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("Top Priority States")
    ws7.sheet_properties.tabColor = "DC143C"

    _write_title(ws7,
                 "States Ranked by C4C Onboarding Need — Highest Volume First",
                 "States sorted by number of accounts NOT yet on C4C. Focus onboarding efforts where the most accounts need registration.",
                 num_cols=8)

    row = 4
    priority_headers = ["Rank", "State", "Not on C4C", "C4C Matched",
                        "Total Installers", "Gap %", "C4C Rate", "Rack Installers"]
    for ci, h in enumerate(priority_headers, 1):
        ws7.cell(row=row, column=ci, value=h)
    _apply_header_row(ws7, row, len(priority_headers))
    row += 1

    state_priority = []
    for sc in sorted(state_data.keys()):
        d = state_data[sc]
        promo = d.get("Promo Only (Not on C4C)", 0)
        both = d.get("On Both Lists", 0)
        c4c = d.get("C4C Only", 0)
        rack = d.get("Rack Installer", 0)
        inst_total = promo + both + c4c
        if inst_total == 0:
            continue
        gap = promo / inst_total * 100
        c4c_rate = (both + c4c) / inst_total * 100
        sname = US_STATE_NAMES.get(sc, sc)
        state_priority.append((sname, sc, promo, both + c4c, inst_total, gap, c4c_rate, rack))

    state_priority.sort(key=lambda x: -x[2])

    for rank, (sname, sc, nc, mc, total, gap, c4c_rate, rack) in enumerate(state_priority, 1):
        ws7.cell(row=row, column=1, value=rank)
        ws7.cell(row=row, column=2, value=f"{sname} ({sc})")
        ws7.cell(row=row, column=3, value=nc)
        ws7.cell(row=row, column=4, value=mc)
        ws7.cell(row=row, column=5, value=total)
        ws7.cell(row=row, column=6, value=f"{gap:.1f}%")
        ws7.cell(row=row, column=7, value=f"{c4c_rate:.1f}%")
        ws7.cell(row=row, column=8, value=rack)

        _apply_data_row(ws7, row, len(priority_headers), alt=(row % 2 == 0))
        for col in [1, 3, 4, 5, 6, 7, 8]:
            ws7.cell(row=row, column=col).alignment = Alignment(horizontal="center")

        if rank <= 10:
            ws7.cell(row=row, column=3).fill = AMBER_FILL
        if gap >= 80:
            ws7.cell(row=row, column=6).fill = RED_FILL
        elif gap >= 50:
            ws7.cell(row=row, column=6).fill = AMBER_FILL
        if c4c_rate >= 50:
            ws7.cell(row=row, column=7).fill = GREEN_FILL
        row += 1

    _auto_width(ws7, len(priority_headers))
    _add_auto_filter(ws7, 4, len(priority_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 8: TOP PRIORITY COUNTIES
    # ═══════════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("Top Priority Counties")
    ws8.sheet_properties.tabColor = "B91C1C"

    _write_title(ws8,
                 "Top 200 Counties by C4C Onboarding Need — Drill-Down Priority",
                 "Counties sorted by number of accounts NOT yet on C4C. Identify specific markets for targeted outreach.",
                 num_cols=9)

    row = 4
    cpri_headers = ["Rank", "County", "State", "Not on C4C", "C4C Matched",
                    "Total Installers", "Gap %", "C4C Rate", "Distributors"]
    for ci, h in enumerate(cpri_headers, 1):
        ws8.cell(row=row, column=ci, value=h)
    _apply_header_row(ws8, row, len(cpri_headers))
    row += 1

    county_priority = []
    for (st, county), d in county_data.items():
        promo = d.get("Promo Only (Not on C4C)", 0)
        both = d.get("On Both Lists", 0)
        c4c_only = d.get("C4C Only", 0)
        dist = d.get("Distributor", 0)
        inst = promo + both + c4c_only
        if inst == 0:
            continue
        gap = promo / inst * 100
        c4c_rate = (both + c4c_only) / inst * 100
        county_priority.append((county, US_STATE_NAMES.get(st, st), st, promo, both + c4c_only, inst, gap, c4c_rate, dist))

    county_priority.sort(key=lambda x: -x[3])

    for rank, (county, sname, sc, nc, mc, total, gap, c4c_rate, dist) in enumerate(county_priority[:200], 1):
        ws8.cell(row=row, column=1, value=rank)
        ws8.cell(row=row, column=2, value=county)
        ws8.cell(row=row, column=3, value=f"{sname} ({sc})")
        ws8.cell(row=row, column=4, value=nc)
        ws8.cell(row=row, column=5, value=mc)
        ws8.cell(row=row, column=6, value=total)
        ws8.cell(row=row, column=7, value=f"{gap:.1f}%")
        ws8.cell(row=row, column=8, value=f"{c4c_rate:.1f}%")
        ws8.cell(row=row, column=9, value=dist if dist else "")

        _apply_data_row(ws8, row, len(cpri_headers), alt=(row % 2 == 0))
        for col in [1, 4, 5, 6, 7, 8, 9]:
            ws8.cell(row=row, column=col).alignment = Alignment(horizontal="center")

        if rank <= 25:
            ws8.cell(row=row, column=4).fill = AMBER_FILL
        if gap >= 80:
            ws8.cell(row=row, column=7).fill = RED_FILL
        elif gap >= 50:
            ws8.cell(row=row, column=7).fill = AMBER_FILL
        if c4c_rate >= 50:
            ws8.cell(row=row, column=8).fill = GREEN_FILL
        if dist:
            ws8.cell(row=row, column=9).fill = GOLD_FILL
        row += 1

    _auto_width(ws8, len(cpri_headers))
    _add_auto_filter(ws8, 4, len(cpri_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 9: DISTRIBUTORS
    # ═══════════════════════════════════════════════════════════════
    ws9 = wb.create_sheet("Distributors")
    ws9.sheet_properties.tabColor = "D97706"

    _write_title(ws9,
                 f"Butler Performance Distributor Directory — {len(distributors)} Locations",
                 "Complete list of authorized RP distributors. Filter by state or county to find nearby distribution.",
                 num_cols=8)

    row = 4
    dist_headers = ["Distributor Name", "Address", "City", "State", "County", "Zip",
                    "Latitude", "Longitude"]
    for ci, h in enumerate(dist_headers, 1):
        ws9.cell(row=row, column=ci, value=h)
    _apply_header_row(ws9, row, len(dist_headers))
    row += 1

    for d in sorted(distributors, key=lambda x: (x.get("state", ""), x.get("store_name", x.get("name", "")))):
        ws9.cell(row=row, column=1, value=d.get("store_name", d.get("name", "")))
        ws9.cell(row=row, column=2, value=d.get("address", ""))
        ws9.cell(row=row, column=3, value=d.get("city", ""))
        ws9.cell(row=row, column=4, value=d.get("state", ""))
        ws9.cell(row=row, column=5, value=d.get("county", ""))
        ws9.cell(row=row, column=6, value=d.get("zip", ""))
        ws9.cell(row=row, column=7, value=d.get("latitude", ""))
        ws9.cell(row=row, column=8, value=d.get("longitude", ""))
        _apply_data_row(ws9, row, len(dist_headers), alt=(row % 2 == 0))
        ws9.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        for col in range(1, len(dist_headers) + 1):
            ws9.cell(row=row, column=col).fill = GOLD_FILL
        row += 1

    _auto_width(ws9, len(dist_headers))
    _add_auto_filter(ws9, 4, len(dist_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 10: RACK INSTALLERS
    # ═══════════════════════════════════════════════════════════════
    ws10 = wb.create_sheet("Rack Installers")
    ws10.sheet_properties.tabColor = "7C3AED"

    all_rack = [c for c in all_accounts if c.get("rack_installer")]
    new_rack = [c for c in all_accounts if c.get("type") == "Rack Installer"]

    _write_title(ws10,
                 f"RP Display Rack Accounts — {len(all_rack)} Flagged ({len(new_rack)} unmatched new)",
                 "All accounts identified as having Butler Performance display racks. Includes matched and unmatched installers.",
                 num_cols=11)

    rack_headers = ["Store Name", "Address", "City", "State", "County", "Zip",
                    "Account Type", "C4C Status", "Country", "Latitude", "Longitude"]
    rack_sorted = sorted(all_rack, key=lambda x: (x.get("state", ""), x.get("county", ""), x.get("store_name", "")))
    row = 4
    for ci, h in enumerate(rack_headers, 1):
        ws10.cell(row=row, column=ci, value=h)
    _apply_header_row(ws10, row, len(rack_headers))
    row += 1
    for c in rack_sorted:
        t = c.get("type", "")
        if t in ("On Both Lists", "C4C Only"):
            c4c_status = "On C4C"
        elif t == "Promo Only (Not on C4C)":
            c4c_status = "NOT on C4C"
        else:
            c4c_status = "Unmatched"
        ws10.cell(row=row, column=1, value=c.get("store_name", ""))
        ws10.cell(row=row, column=2, value=c.get("address", ""))
        ws10.cell(row=row, column=3, value=c.get("city", ""))
        ws10.cell(row=row, column=4, value=c.get("state", ""))
        ws10.cell(row=row, column=5, value=c.get("county", ""))
        ws10.cell(row=row, column=6, value=c.get("zip", ""))
        ws10.cell(row=row, column=7, value=t)
        ws10.cell(row=row, column=8, value=c4c_status)
        ws10.cell(row=row, column=9, value=c.get("country", "US"))
        ws10.cell(row=row, column=10, value=c.get("latitude", ""))
        ws10.cell(row=row, column=11, value=c.get("longitude", ""))
        _apply_data_row(ws10, row, len(rack_headers), alt=(row % 2 == 0))
        ws10.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        tfill = _type_fill(t)
        if tfill:
            ws10.cell(row=row, column=7).fill = tfill
        if c4c_status == "On C4C":
            ws10.cell(row=row, column=8).fill = GREEN_FILL
        elif c4c_status == "NOT on C4C":
            ws10.cell(row=row, column=8).fill = RED_FILL
        else:
            ws10.cell(row=row, column=8).fill = AMBER_FILL
        row += 1
    _auto_width(ws10, len(rack_headers))
    _add_auto_filter(ws10, 4, len(rack_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 11: POWERSPORTS / MOTORSPORTS
    # ═══════════════════════════════════════════════════════════════
    ws11 = wb.create_sheet("Powersports")
    ws11.sheet_properties.tabColor = "E11D48"

    _write_title(ws11,
                 f"Powersports / Motorsports Accounts — {len(powersports)} Locations",
                 "Complete directory of powersports and motorsports partners.",
                 num_cols=9)

    ps_headers = ["Store Name", "Address", "City", "State", "County", "Zip",
                  "Country", "Latitude", "Longitude"]
    ps_sorted = sorted(powersports, key=lambda x: (x.get("state", ""), x.get("city", ""), x.get("store_name", "")))
    row = 4
    for ci, h in enumerate(ps_headers, 1):
        ws11.cell(row=row, column=ci, value=h)
    _apply_header_row(ws11, row, len(ps_headers))
    row += 1
    for c in ps_sorted:
        ws11.cell(row=row, column=1, value=c.get("store_name", ""))
        ws11.cell(row=row, column=2, value=c.get("address", ""))
        ws11.cell(row=row, column=3, value=c.get("city", ""))
        ws11.cell(row=row, column=4, value=c.get("state", ""))
        ws11.cell(row=row, column=5, value=c.get("county", ""))
        ws11.cell(row=row, column=6, value=c.get("zip", ""))
        ws11.cell(row=row, column=7, value=c.get("country", "US"))
        ws11.cell(row=row, column=8, value=c.get("latitude", ""))
        ws11.cell(row=row, column=9, value=c.get("longitude", ""))
        _apply_data_row(ws11, row, len(ps_headers), alt=(row % 2 == 0))
        ws11.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        row += 1
    _auto_width(ws11, len(ps_headers))
    _add_auto_filter(ws11, 4, len(ps_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 12: INTERNATIONAL
    # ═══════════════════════════════════════════════════════════════
    ws12 = wb.create_sheet("International")
    ws12.sheet_properties.tabColor = "4F46E5"

    _write_title(ws12,
                 f"International Accounts — {len(international)} Locations",
                 "Global Butler Performance partners outside the US and Canada.",
                 num_cols=9)

    intl_headers = ["Store Name", "Address", "City", "State/Region", "Country", "Zip",
                    "Latitude", "Longitude"]
    intl_sorted = sorted(international, key=lambda x: (x.get("country", ""), x.get("store_name", "")))
    row = 4
    for ci, h in enumerate(intl_headers, 1):
        ws12.cell(row=row, column=ci, value=h)
    _apply_header_row(ws12, row, len(intl_headers))
    row += 1
    for c in intl_sorted:
        ws12.cell(row=row, column=1, value=c.get("store_name", ""))
        ws12.cell(row=row, column=2, value=c.get("address", ""))
        ws12.cell(row=row, column=3, value=c.get("city", ""))
        ws12.cell(row=row, column=4, value=c.get("state", ""))
        ws12.cell(row=row, column=5, value=c.get("country", ""))
        ws12.cell(row=row, column=6, value=c.get("zip", ""))
        ws12.cell(row=row, column=7, value=c.get("latitude", ""))
        ws12.cell(row=row, column=8, value=c.get("longitude", ""))
        _apply_data_row(ws12, row, len(intl_headers), alt=(row % 2 == 0))
        row += 1
    _auto_width(ws12, len(intl_headers))
    _add_auto_filter(ws12, 4, len(intl_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 13: CANADA
    # ═══════════════════════════════════════════════════════════════
    ws13 = wb.create_sheet("Canada")
    ws13.sheet_properties.tabColor = "059669"

    _write_title(ws13,
                 f"Canadian Accounts — {len(canada)} Locations",
                 "Butler Performance partners in Canada.",
                 num_cols=9)

    ca_headers = ["Store Name", "Address", "City", "Province", "Zip/Postal",
                  "Country", "Latitude", "Longitude"]
    ca_sorted = sorted(canada, key=lambda x: (x.get("state", ""), x.get("store_name", "")))
    row = 4
    for ci, h in enumerate(ca_headers, 1):
        ws13.cell(row=row, column=ci, value=h)
    _apply_header_row(ws13, row, len(ca_headers))
    row += 1
    for c in ca_sorted:
        ws13.cell(row=row, column=1, value=c.get("store_name", ""))
        ws13.cell(row=row, column=2, value=c.get("address", ""))
        ws13.cell(row=row, column=3, value=c.get("city", ""))
        ws13.cell(row=row, column=4, value=c.get("state", ""))
        ws13.cell(row=row, column=5, value=c.get("zip", ""))
        ws13.cell(row=row, column=6, value=c.get("country", "CA"))
        ws13.cell(row=row, column=7, value=c.get("latitude", ""))
        ws13.cell(row=row, column=8, value=c.get("longitude", ""))
        _apply_data_row(ws13, row, len(ca_headers), alt=(row % 2 == 0))
        row += 1
    _auto_width(ws13, len(ca_headers))
    _add_auto_filter(ws13, 4, len(ca_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 14: DISTRIBUTOR COVERAGE ANALYSIS
    # ═══════════════════════════════════════════════════════════════
    ws14 = wb.create_sheet("Distributor Coverage")
    ws14.sheet_properties.tabColor = "B45309"

    _write_title(ws14,
                 "Distributor Coverage vs. Installer Density",
                 "States ranked by installer count, showing whether distributor coverage exists. Identifies distribution gaps.",
                 num_cols=7)

    row = 4
    dc_headers = ["State", "Code", "Installer Accounts", "Distributors",
                  "Installers per Distributor", "Has Coverage", "Coverage Gap"]
    for ci, h in enumerate(dc_headers, 1):
        ws14.cell(row=row, column=ci, value=h)
    _apply_header_row(ws14, row, len(dc_headers))
    row += 1

    dist_by_state = Counter()
    for d in distributors:
        st = d.get("state", "")
        if st:
            dist_by_state[st] += 1

    coverage_data = []
    for sc in sorted(state_data.keys()):
        d = state_data[sc]
        promo = d.get("Promo Only (Not on C4C)", 0)
        both = d.get("On Both Lists", 0)
        c4c = d.get("C4C Only", 0)
        rack = d.get("Rack Installer", 0)
        inst_count = promo + both + c4c + rack
        dist_count = dist_by_state.get(sc, 0)
        ratio = inst_count / dist_count if dist_count else 0
        has_coverage = "Yes" if dist_count > 0 else "No"
        gap = "Covered" if dist_count > 0 else ("Gap" if inst_count > 0 else "No installers")
        coverage_data.append((US_STATE_NAMES.get(sc, sc), sc, inst_count, dist_count, ratio, has_coverage, gap))

    coverage_data.sort(key=lambda x: -x[2])

    for (sname, sc, inst, dist, ratio, has_cov, gap) in coverage_data:
        ws14.cell(row=row, column=1, value=sname)
        ws14.cell(row=row, column=2, value=sc)
        ws14.cell(row=row, column=3, value=inst)
        ws14.cell(row=row, column=4, value=dist)
        ws14.cell(row=row, column=5, value=f"{ratio:.0f}:1" if dist else "No dist.")
        ws14.cell(row=row, column=6, value=has_cov)
        ws14.cell(row=row, column=7, value=gap)

        _apply_data_row(ws14, row, len(dc_headers), alt=(row % 2 == 0))
        for col in range(2, 8):
            ws14.cell(row=row, column=col).alignment = Alignment(horizontal="center")

        if has_cov == "No" and inst > 0:
            ws14.cell(row=row, column=7).fill = RED_FILL
            ws14.cell(row=row, column=7).font = Font(name="Calibri", bold=True, size=10, color="DC2626")
        elif has_cov == "Yes":
            ws14.cell(row=row, column=6).fill = GREEN_FILL
            if dist > 0:
                ws14.cell(row=row, column=4).fill = GOLD_FILL
        row += 1

    _auto_width(ws14, len(dc_headers))
    _add_auto_filter(ws14, 4, len(dc_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 15: RECONCILIATION
    # ═══════════════════════════════════════════════════════════════
    xref = _cross_analyze_lists()

    ws15 = wb.create_sheet("Reconciliation")
    ws15.sheet_properties.tabColor = "2563EB"

    _write_title(ws15,
                 "C4C vs Promo List — Cross-Reference & Reconciliation",
                 "Detailed comparison of source Excel files: C4C Installer List vs. Promo Participation List.",
                 num_cols=4)

    row = 4
    ws15.merge_cells(f"A{row}:B{row}")
    ws15[f"A{row}"] = "Source File Totals"
    ws15[f"A{row}"].font = SECTION_FONT
    row += 1

    for ci, h in enumerate(["Metric", "Count"], 1):
        ws15.cell(row=row, column=ci, value=h)
    _apply_header_row(ws15, row, 2)
    row += 1

    totals = [
        ("C4C Installer List — Total Rows", xref["c4c_rows"]),
        ("C4C Installer List — Unique Names", xref["c4c_unique"]),
        ("Promo Participation List — Total Rows", xref["promo_rows"]),
        ("Promo Participation List — Unique Names", xref["promo_unique"]),
        ("", ""),
        ("Matched (name on BOTH lists)", xref["matched"]),
        ("On C4C ONLY (not on Promo list)", xref["only_c4c"]),
        ("On Promo ONLY (not on C4C)", xref["only_promo"]),
        ("", ""),
        ("C4C Internal Duplicates (same name, multiple rows)", len(xref["c4c_dupes"])),
        ("Promo Internal Duplicates (same name, multiple rows)", len(xref["promo_dupes"])),
    ]

    for label, val in totals:
        ws15.cell(row=row, column=1, value=label)
        ws15.cell(row=row, column=2, value=val if val != "" else "")
        _apply_data_row(ws15, row, 2, alt=(row % 2 == 0))
        ws15.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        if label.startswith("Matched") or label.startswith("On C4C") or label.startswith("On Promo"):
            ws15.cell(row=row, column=1).font = BOLD_FONT
            ws15.cell(row=row, column=2).font = BOLD_FONT
        row += 1

    _auto_width(ws15, 2)
    ws15.column_dimensions["A"].width = 55

    # ═══════════════════════════════════════════════════════════════
    # SHEET 16: C4C DUPLICATES
    # ═══════════════════════════════════════════════════════════════
    ws16 = wb.create_sheet("C4C Duplicates")
    ws16.sheet_properties.tabColor = "DC143C"

    _write_title(ws16,
                 f"C4C Installer List — Internal Duplicates ({len(xref['c4c_dupes'])} names)",
                 "Names appearing multiple times in the C4C Installer List source file.",
                 num_cols=3)

    row = 4
    dupe_headers = ["Installer Name", "Occurrences", "Locations"]
    for ci, h in enumerate(dupe_headers, 1):
        ws16.cell(row=row, column=ci, value=h)
    _apply_header_row(ws16, row, len(dupe_headers))
    row += 1

    for d in xref["c4c_dupes"]:
        ws16.cell(row=row, column=1, value=d["name"])
        ws16.cell(row=row, column=2, value=d["count"])
        ws16.cell(row=row, column=3, value=d["locations"])
        _apply_data_row(ws16, row, len(dupe_headers), alt=(row % 2 == 0))
        ws16.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1

    _auto_width(ws16, len(dupe_headers))
    ws16.column_dimensions["C"].width = 80
    _add_auto_filter(ws16, 4, len(dupe_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 17: PROMO DUPLICATES
    # ═══════════════════════════════════════════════════════════════
    ws17 = wb.create_sheet("Promo Duplicates")
    ws17.sheet_properties.tabColor = "D97706"

    _write_title(ws17,
                 f"Promo List — Internal Duplicates ({len(xref['promo_dupes'])} names)",
                 "Names appearing multiple times in the Promo Participation source file.",
                 num_cols=3)

    row = 4
    for ci, h in enumerate(dupe_headers, 1):
        ws17.cell(row=row, column=ci, value=h)
    _apply_header_row(ws17, row, len(dupe_headers))
    row += 1

    for d in xref["promo_dupes"]:
        ws17.cell(row=row, column=1, value=d["name"])
        ws17.cell(row=row, column=2, value=d["count"])
        ws17.cell(row=row, column=3, value=d["locations"])
        _apply_data_row(ws17, row, len(dupe_headers), alt=(row % 2 == 0))
        ws17.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1

    _auto_width(ws17, len(dupe_headers))
    ws17.column_dimensions["C"].width = 80
    _add_auto_filter(ws17, 4, len(dupe_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 18: ON C4C ONLY (not on Promo)
    # ═══════════════════════════════════════════════════════════════
    ws18 = wb.create_sheet("On C4C Only")
    ws18.sheet_properties.tabColor = "e31837"

    _write_title(ws18,
                 f"Accounts on C4C but NOT on Promo List — {xref['only_c4c']} Accounts",
                 "These accounts are registered in C4C but do not appear on the Promotion Participation list.",
                 num_cols=4)

    row = 4
    c4c_only_headers = ["Installer Name", "City", "State", "Account ID"]
    for ci, h in enumerate(c4c_only_headers, 1):
        ws18.cell(row=row, column=ci, value=h)
    _apply_header_row(ws18, row, len(c4c_only_headers))
    row += 1

    for entry in xref["only_c4c_names"]:
        ws18.cell(row=row, column=1, value=entry["name"])
        ws18.cell(row=row, column=2, value=entry["city"])
        ws18.cell(row=row, column=3, value=entry["state"])
        ws18.cell(row=row, column=4, value=entry["acct_id"])
        _apply_data_row(ws18, row, len(c4c_only_headers), alt=(row % 2 == 0))
        ws18.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        row += 1

    _auto_width(ws18, len(c4c_only_headers))
    _add_auto_filter(ws18, 4, len(c4c_only_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 19: FAILED TO GEOLOCATE
    # ═══════════════════════════════════════════════════════════════
    failed = _get_failed_geolocations()

    ws19 = wb.create_sheet("Failed to Geolocate")
    ws19.sheet_properties.tabColor = "94A3B8"

    _write_title(ws19,
                 f"Failed to Geolocate — {len(failed)} Accounts",
                 "These accounts could not be placed on the map due to invalid or missing address data.",
                 num_cols=9)

    row = 4
    fail_headers = ["Store Name", "Address", "City", "State", "Zip",
                     "Phone", "Email", "Source Sheet", "Reason"]
    for ci, h in enumerate(fail_headers, 1):
        ws19.cell(row=row, column=ci, value=h)
    _apply_header_row(ws19, row, len(fail_headers))
    row += 1

    for f in sorted(failed, key=lambda x: (x["reason"], x["state"], x["store_name"])):
        ws19.cell(row=row, column=1, value=f["store_name"])
        ws19.cell(row=row, column=2, value=f["address"])
        ws19.cell(row=row, column=3, value=f["city"])
        ws19.cell(row=row, column=4, value=f["state"])
        ws19.cell(row=row, column=5, value=f["zip"])
        ws19.cell(row=row, column=6, value=f["phone"])
        ws19.cell(row=row, column=7, value=f["email"])
        ws19.cell(row=row, column=8, value=f["source"])
        ws19.cell(row=row, column=9, value=f["reason"])

        _apply_data_row(ws19, row, len(fail_headers), alt=(row % 2 == 0))
        ws19.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        row += 1

    _auto_width(ws19, len(fail_headers))
    _add_auto_filter(ws19, 4, len(fail_headers))

    # ═══════════════════════════════════════════════════════════════
    # SHEET 20: RPO AUTOCARE — ALL ACCOUNTS
    # ═══════════════════════════════════════════════════════════════
    RPO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rpo_autocare_processed.json")
    rpo_data = []
    if os.path.exists(RPO_PATH):
        with open(RPO_PATH) as _rpf:
            rpo_data = json.load(_rpf)

    if rpo_data:
        ORANGE_FILL = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")

        ws20 = wb.create_sheet("RPO Autocare — All")
        ws20.sheet_properties.tabColor = "F97316"
        ws20.freeze_panes = "A5"

        ws20.merge_cells("A1:H1")
        ws20["A1"] = "RPO Autocare 2025 — All Accounts with C4C Status"
        ws20["A1"].font = TITLE_FONT
        ws20.row_dimensions[1].height = 28

        rpo_on_c4c = sum(1 for a in rpo_data if a.get("c4c_status") == "On C4C")
        rpo_not_c4c = len(rpo_data) - rpo_on_c4c
        rpo_total_sales = sum(a.get("cytd_sales", 0) for a in rpo_data)
        rpo_not_c4c_sales = sum(a.get("cytd_sales", 0) for a in rpo_data if a.get("c4c_status") != "On C4C")

        ws20.merge_cells("A2:H2")
        ws20["A2"] = (
            f"{len(rpo_data):,} total accounts  |  "
            f"{rpo_on_c4c:,} on C4C ({rpo_on_c4c/len(rpo_data)*100:.1f}%)  |  "
            f"{rpo_not_c4c:,} NOT on C4C  |  "
            f"Non-C4C CYTD Revenue: ${rpo_not_c4c_sales:,.0f}"
        )
        ws20["A2"].font = Font(name="Calibri", italic=True, size=10, color="64748B")
        ws20.row_dimensions[2].height = 20
        ws20.row_dimensions[3].height = 6

        rpo_all_headers = [
            "Installer Name", "C4C Status", "CYTD Sales", "Gold Flag",
            "District", "Region", "Company Owned", "City",
        ]

        for ci, h in enumerate(rpo_all_headers, 1):
            cell = ws20.cell(row=4, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        rpo_sorted = sorted(rpo_data, key=lambda x: x.get("name", "").upper())
        row = 5
        for a in rpo_sorted:
            ws20.cell(row=row, column=1, value=a.get("name", ""))
            c4c_cell = ws20.cell(row=row, column=2, value=a.get("c4c_status", ""))
            ws20.cell(row=row, column=3, value=a.get("cytd_sales", 0))
            ws20.cell(row=row, column=4, value=a.get("gold_flag", ""))
            ws20.cell(row=row, column=5, value=a.get("district", ""))
            ws20.cell(row=row, column=6, value=a.get("region", ""))
            ws20.cell(row=row, column=7, value=a.get("company_owned", ""))
            ws20.cell(row=row, column=8, value=a.get("city", ""))

            _apply_data_row(ws20, row, len(rpo_all_headers), alt=(row % 2 == 0))
            ws20.cell(row=row, column=3).number_format = '$#,##0.00'
            ws20.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            ws20.cell(row=row, column=7).alignment = Alignment(horizontal="center")

            status = a.get("c4c_status", "")
            if status == "On C4C":
                c4c_cell.fill = GREEN_FILL
            elif status == "Promo Only":
                c4c_cell.fill = AMBER_FILL
            elif status == "Rack Only":
                c4c_cell.fill = PURPLE_LIGHT_FILL
            else:
                c4c_cell.fill = RED_FILL

            row += 1

        _auto_width(ws20, len(rpo_all_headers))
        _add_auto_filter(ws20, 4, len(rpo_all_headers))

        # ═══════════════════════════════════════════════════════════════
        # SHEET 21: RPO AUTOCARE — NOT ON C4C (by CYTD Sales)
        # ═══════════════════════════════════════════════════════════════
        ws21 = wb.create_sheet("RPO — Not on C4C")
        ws21.sheet_properties.tabColor = "DC2626"
        ws21.freeze_panes = "A5"

        ws21.merge_cells("A1:H1")
        ws21["A1"] = "RPO Autocare 2025 — Accounts NOT on C4C (Prioritized by Sales)"
        ws21["A1"].font = TITLE_FONT
        ws21.row_dimensions[1].height = 28

        ws21.merge_cells("A2:H2")
        ws21["A2"] = (
            f"{rpo_not_c4c:,} accounts not on C4C  |  "
            f"CYTD Revenue at Risk: ${rpo_not_c4c_sales:,.0f}  |  "
            f"Sorted by sales volume — highest priority targets at top"
        )
        ws21["A2"].font = Font(name="Calibri", italic=True, size=10, color="64748B")
        ws21.row_dimensions[2].height = 20
        ws21.row_dimensions[3].height = 6

        rpo_notc4c_headers = [
            "Installer Name", "C4C Status", "CYTD Sales", "Gold Flag",
            "District", "Region", "Company Owned", "City",
        ]

        for ci, h in enumerate(rpo_notc4c_headers, 1):
            cell = ws21.cell(row=4, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        rpo_not_c4c_list = sorted(
            [a for a in rpo_data if a.get("c4c_status") != "On C4C"],
            key=lambda x: -x.get("cytd_sales", 0),
        )
        row = 5
        for a in rpo_not_c4c_list:
            ws21.cell(row=row, column=1, value=a.get("name", ""))
            c4c_cell = ws21.cell(row=row, column=2, value=a.get("c4c_status", ""))
            ws21.cell(row=row, column=3, value=a.get("cytd_sales", 0))
            ws21.cell(row=row, column=4, value=a.get("gold_flag", ""))
            ws21.cell(row=row, column=5, value=a.get("district", ""))
            ws21.cell(row=row, column=6, value=a.get("region", ""))
            ws21.cell(row=row, column=7, value=a.get("company_owned", ""))
            ws21.cell(row=row, column=8, value=a.get("city", ""))

            _apply_data_row(ws21, row, len(rpo_notc4c_headers), alt=(row % 2 == 0))
            ws21.cell(row=row, column=3).number_format = '$#,##0.00'
            ws21.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            ws21.cell(row=row, column=7).alignment = Alignment(horizontal="center")

            status = a.get("c4c_status", "")
            if status == "Promo Only":
                c4c_cell.fill = AMBER_FILL
            elif status == "Rack Only":
                c4c_cell.fill = PURPLE_LIGHT_FILL
            else:
                c4c_cell.fill = RED_FILL

            row += 1

        _auto_width(ws21, len(rpo_notc4c_headers))
        _add_auto_filter(ws21, 4, len(rpo_notc4c_headers))

    # ═══════════════════════════════════════════════════════════════
    # SAVE
    # ═══════════════════════════════════════════════════════════════
    wb.save(output_path)

    sheet_count = len(wb.sheetnames)

    return {
        "total_accounts": len(all_accounts),
        "total_installers": total_installers,
        "not_on_c4c": len(not_c4c),
        "c4c_matched": len(c4c_matched),
        "states": len(us_states),
        "counties": len(county_set),
        "distributors": len(distributors),
        "powersports": len(powersports),
        "international": len(international),
        "canada": len(canada),
        "rack_flagged": rack_flagged,
        "failed_geo": len(failed),
        "c4c_dupes": len(xref["c4c_dupes"]),
        "promo_dupes": len(xref["promo_dupes"]),
        "cross_matched": xref["matched"],
        "rpo_total": len(rpo_data),
        "rpo_not_c4c": len([a for a in rpo_data if a.get("c4c_status") != "On C4C"]) if rpo_data else 0,
        "sheets": sheet_count,
    }
