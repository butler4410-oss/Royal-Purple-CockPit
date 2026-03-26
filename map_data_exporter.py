import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from collections import Counter

CUSTOMERS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "customers.json")

PURPLE = "1a1a2e"
PURPLE_MED = "e31837"
GOLD = "e31837"
RED_ACCENT = "DC2626"
GREEN_ACCENT = "16A34A"
BLUE_ACCENT = "2563EB"

HEADER_FILL = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color=PURPLE_MED, end_color=PURPLE_MED, fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color=PURPLE)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color=PURPLE_MED)
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color=PURPLE_MED)
AMBER_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FDF6E3", end_color="FDF6E3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color=PURPLE))

US_STATE_NAMES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "DC": "District of Columbia", "FL": "Florida", "GA": "Georgia", "HI": "Hawaii",
    "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine",
    "MD": "Maryland", "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota",
    "MS": "Mississippi", "MO": "Missouri", "MT": "Montana", "NE": "Nebraska",
    "NV": "Nevada", "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico",
    "NY": "New York", "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio",
    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "PR": "Puerto Rico",
    "RI": "Rhode Island", "SC": "South Carolina", "SD": "South Dakota",
    "TN": "Tennessee", "TX": "Texas", "UT": "Utah", "VT": "Vermont",
    "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming",
}


def _apply_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 30


def _apply_data_row(ws, row, num_cols, alt=False):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        if alt:
            cell.fill = LIGHT_GRAY


def _auto_width(ws, num_cols, max_width=42):
    for ci in range(1, num_cols + 1):
        letter = get_column_letter(ci)
        max_len = 10
        for row in ws.iter_rows(min_col=ci, max_col=ci, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 3, max_width)


def _add_autofilter(ws, header_row, num_cols, last_row):
    start = f"A{header_row}"
    end = f"{get_column_letter(num_cols)}{last_row}"
    ws.auto_filter.ref = f"{start}:{end}"


AMBER_DARK_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
PURPLE_LIGHT_FILL = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
ROSE_FILL = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
INDIGO_FILL = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
EMERALD_FILL = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

def _type_fill(account_type):
    if "Promo Only" in account_type:
        return RED_FILL
    elif "Both" in account_type:
        return GREEN_FILL
    elif "C4C Only" in account_type:
        return BLUE_FILL
    elif "Rack Installer" in account_type:
        return PURPLE_LIGHT_FILL
    elif "Distributor" in account_type:
        return AMBER_DARK_FILL
    elif "Powersports" in account_type:
        return ROSE_FILL
    elif "International" in account_type:
        return INDIGO_FILL
    elif "Canada" in account_type:
        return EMERALD_FILL
    return None


def generate_map_export(output_path, customers=None):
    if customers is None:
        if os.path.exists(CUSTOMERS_PATH):
            with open(CUSTOMERS_PATH, "r") as f:
                customers = json.load(f)
        else:
            customers = []

    customers = sorted(customers, key=lambda x: (
        x.get("state", ""), x.get("county", ""), x.get("store_name", "")
    ))

    non_installer_types = {"Distributor", "Powersports/Motorsports", "International", "Canada"}
    installers = [c for c in customers if c.get("type") not in non_installer_types]

    by_state = {}
    for c in installers:
        st = c.get("state", "Other")
        if st not in by_state:
            by_state[st] = []
        by_state[st].append(c)

    type_counts = Counter(c.get("type", "Unknown") for c in customers)
    county_counts = Counter(c.get("county", "") for c in installers if c.get("county"))
    state_counties = {}
    for c in installers:
        st = c.get("state", "")
        co = c.get("county", "")
        if st and co:
            if st not in state_counties:
                state_counties[st] = set()
            state_counties[st].add(co)

    wb = Workbook()

    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_properties.tabColor = PURPLE

    ws_dash.merge_cells("A1:H1")
    ws_dash["A1"] = "Butler Performance Installer Account Map Data"
    ws_dash["A1"].font = Font(name="Calibri", bold=True, size=18, color=PURPLE)
    ws_dash.row_dimensions[1].height = 36

    ws_dash.merge_cells("A2:H2")
    ws_dash["A2"] = "Complete account listing by state and county with C4C status"
    ws_dash["A2"].font = Font(name="Calibri", size=11, color="64748B")

    row = 4
    ws_dash.merge_cells(f"A{row}:C{row}")
    ws_dash[f"A{row}"] = "Account Summary"
    ws_dash[f"A{row}"].font = SUBTITLE_FONT
    for c in range(1, 4):
        ws_dash.cell(row=row, column=c).border = BOTTOM_BORDER
    row += 1

    distributor_count = type_counts.get("Distributor", 0)
    ps_count = type_counts.get("Powersports/Motorsports", 0)
    intl_count = type_counts.get("International", 0)
    ca_count = type_counts.get("Canada", 0)
    rack_count = type_counts.get("Rack Installer", 0)
    installer_count = (type_counts.get("Promo Only (Not on C4C)", 0) +
                       type_counts.get("On Both Lists", 0) +
                       type_counts.get("C4C Only", 0) +
                       rack_count)

    metrics = [
        ("Total Locations", len(customers)),
        ("Installer Accounts", installer_count),
        ("Distributors", distributor_count),
        ("Powersports/Motorsports", ps_count),
        ("International", intl_count),
        ("Canada", ca_count),
        ("States / Regions", len(by_state)),
        ("Counties (US)", len(county_counts)),
        ("", ""),
        ("Promo Only (Not on C4C)", type_counts.get("Promo Only (Not on C4C)", 0)),
        ("On Both Lists", type_counts.get("On Both Lists", 0)),
        ("C4C Only", type_counts.get("C4C Only", 0)),
        ("Rack Installer", rack_count),
    ]

    for label, val in metrics:
        if not label:
            row += 1
            continue
        ws_dash.cell(row=row, column=1, value=label).font = BOLD_FONT
        cell = ws_dash.cell(row=row, column=2, value=val)
        cell.font = Font(name="Calibri", bold=True, size=11, color=PURPLE)
        cell.alignment = Alignment(horizontal="right")

        if "Rack Installer" in label:
            ws_dash.cell(row=row, column=1).fill = PURPLE_LIGHT_FILL
        elif "Distributor" in label:
            ws_dash.cell(row=row, column=1).fill = AMBER_DARK_FILL
        elif "Powersports" in label:
            ws_dash.cell(row=row, column=1).fill = ROSE_FILL
        elif "International" in label:
            ws_dash.cell(row=row, column=1).fill = INDIGO_FILL
        elif "Canada" in label:
            ws_dash.cell(row=row, column=1).fill = EMERALD_FILL
        elif "Promo Only" in label:
            ws_dash.cell(row=row, column=1).fill = RED_FILL
        elif "Both" in label:
            ws_dash.cell(row=row, column=1).fill = GREEN_FILL
        elif "C4C Only" in label:
            ws_dash.cell(row=row, column=1).fill = BLUE_FILL

        row += 1

    row += 2
    ws_dash.merge_cells(f"A{row}:H{row}")
    ws_dash[f"A{row}"] = "State Breakdown"
    ws_dash[f"A{row}"].font = SUBTITLE_FONT
    for c in range(1, 9):
        ws_dash.cell(row=row, column=c).border = BOTTOM_BORDER
    row += 1

    state_headers = ["State", "Code", "Total", "Counties", "Promo Only", "On Both Lists", "C4C Only", "Gap %"]
    for ci, h in enumerate(state_headers, 1):
        ws_dash.cell(row=row, column=ci, value=h)
    _apply_header_row(ws_dash, row, len(state_headers))
    header_row_state = row
    row += 1

    state_data_start = row
    for sc in sorted(by_state.keys()):
        accts = by_state[sc]
        state_name = US_STATE_NAMES.get(sc, sc)
        total = len(accts)
        promo = sum(1 for a in accts if a.get("type") == "Promo Only (Not on C4C)")
        both = sum(1 for a in accts if a.get("type") == "On Both Lists")
        c4c = sum(1 for a in accts if a.get("type") == "C4C Only")
        counties = len(state_counties.get(sc, set()))
        gap = promo / total * 100 if total > 0 else 0

        ws_dash.cell(row=row, column=1, value=state_name)
        ws_dash.cell(row=row, column=2, value=sc)
        ws_dash.cell(row=row, column=3, value=total)
        ws_dash.cell(row=row, column=4, value=counties)
        ws_dash.cell(row=row, column=5, value=promo)
        ws_dash.cell(row=row, column=6, value=both)
        ws_dash.cell(row=row, column=7, value=c4c)
        ws_dash.cell(row=row, column=8, value=f"{gap:.1f}%")

        _apply_data_row(ws_dash, row, len(state_headers), alt=(row % 2 == 0))
        for col in [2, 3, 4, 5, 6, 7, 8]:
            ws_dash.cell(row=row, column=col).alignment = Alignment(horizontal="center")

        if gap >= 80:
            ws_dash.cell(row=row, column=8).fill = RED_FILL
            ws_dash.cell(row=row, column=8).font = Font(name="Calibri", bold=True, size=10, color=RED_ACCENT)
        elif gap >= 50:
            ws_dash.cell(row=row, column=8).fill = AMBER_FILL

        row += 1

    _add_autofilter(ws_dash, header_row_state, len(state_headers), row - 1)

    totals_row = row
    ws_dash.cell(row=totals_row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
    ws_dash.cell(row=totals_row, column=3, value=len(customers)).font = Font(name="Calibri", bold=True, size=11)
    ws_dash.cell(row=totals_row, column=4, value=len(county_counts)).font = Font(name="Calibri", bold=True, size=11)
    ws_dash.cell(row=totals_row, column=5, value=type_counts.get("Promo Only (Not on C4C)", 0)).font = Font(name="Calibri", bold=True, size=11)
    ws_dash.cell(row=totals_row, column=6, value=type_counts.get("On Both Lists", 0)).font = Font(name="Calibri", bold=True, size=11)
    ws_dash.cell(row=totals_row, column=7, value=type_counts.get("C4C Only", 0)).font = Font(name="Calibri", bold=True, size=11)
    for col in range(1, len(state_headers) + 1):
        ws_dash.cell(row=totals_row, column=col).border = Border(top=Side(style="double", color=PURPLE))
        ws_dash.cell(row=totals_row, column=col).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=totals_row, column=1).alignment = Alignment(horizontal="left")

    row += 2
    ws_dash.merge_cells(f"A{row}:H{row}")
    ws_dash[f"A{row}"] = "Top 15 Counties by Account Volume"
    ws_dash[f"A{row}"].font = SUBTITLE_FONT
    for c in range(1, 9):
        ws_dash.cell(row=row, column=c).border = BOTTOM_BORDER
    row += 1

    county_headers = ["Rank", "County", "State", "Total", "Promo Only", "On Both Lists", "C4C Only"]
    for ci, h in enumerate(county_headers, 1):
        ws_dash.cell(row=row, column=ci, value=h)
    _apply_header_row(ws_dash, row, len(county_headers))
    row += 1

    county_details = {}
    for c in installers:
        co = c.get("county", "")
        st = c.get("state", "")
        if not co:
            continue
        key = (co, st)
        if key not in county_details:
            county_details[key] = {"total": 0, "promo": 0, "both": 0, "c4c": 0}
        county_details[key]["total"] += 1
        t = c.get("type", "")
        if "Promo Only" in t:
            county_details[key]["promo"] += 1
        elif "Both" in t:
            county_details[key]["both"] += 1
        elif "C4C Only" in t:
            county_details[key]["c4c"] += 1

    top_counties = sorted(county_details.items(), key=lambda x: -x[1]["total"])[:15]

    for rank, ((county, state), stats) in enumerate(top_counties, 1):
        ws_dash.cell(row=row, column=1, value=rank)
        ws_dash.cell(row=row, column=2, value=county)
        ws_dash.cell(row=row, column=3, value=state)
        ws_dash.cell(row=row, column=4, value=stats["total"])
        ws_dash.cell(row=row, column=5, value=stats["promo"])
        ws_dash.cell(row=row, column=6, value=stats["both"])
        ws_dash.cell(row=row, column=7, value=stats["c4c"])

        _apply_data_row(ws_dash, row, len(county_headers), alt=(row % 2 == 0))
        for col in [1, 3, 4, 5, 6, 7]:
            ws_dash.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        if rank <= 5:
            ws_dash.cell(row=row, column=4).fill = GOLD_FILL
            ws_dash.cell(row=row, column=4).font = Font(name="Calibri", bold=True, size=10)
        row += 1

    _auto_width(ws_dash, 8)
    ws_dash.column_dimensions["A"].width = 28
    ws_dash.sheet_view.showGridLines = False

    ws_all = wb.create_sheet("All Accounts")
    ws_all.sheet_properties.tabColor = "64748B"

    all_headers = ["Store Name", "Address", "City", "State", "County", "Zip",
                    "Country", "Account Type", "Latitude", "Longitude"]
    num_all = len(all_headers)

    ws_all.merge_cells(f"A1:{get_column_letter(num_all)}1")
    ws_all["A1"] = f"All Installer Accounts ({len(customers)} total)"
    ws_all["A1"].font = TITLE_FONT
    ws_all.row_dimensions[1].height = 30

    row = 3
    for ci, h in enumerate(all_headers, 1):
        ws_all.cell(row=row, column=ci, value=h)
    _apply_header_row(ws_all, row, num_all)
    header_row_all = row
    row += 1

    for c in customers:
        ws_all.cell(row=row, column=1, value=c.get("store_name", ""))
        ws_all.cell(row=row, column=2, value=c.get("address", ""))
        ws_all.cell(row=row, column=3, value=c.get("city", ""))
        ws_all.cell(row=row, column=4, value=c.get("state", ""))
        ws_all.cell(row=row, column=5, value=c.get("county", ""))
        ws_all.cell(row=row, column=6, value=c.get("zip", ""))
        ws_all.cell(row=row, column=7, value=c.get("country", "US"))
        acct_type = c.get("type", "")
        ws_all.cell(row=row, column=8, value=acct_type)
        ws_all.cell(row=row, column=9, value=c.get("latitude", ""))
        ws_all.cell(row=row, column=10, value=c.get("longitude", ""))

        _apply_data_row(ws_all, row, num_all, alt=(row % 2 == 0))
        ws_all.cell(row=row, column=4).alignment = Alignment(horizontal="center")

        tf = _type_fill(acct_type)
        if tf:
            ws_all.cell(row=row, column=8).fill = tf

        row += 1

    _add_autofilter(ws_all, header_row_all, num_all, row - 1)
    _auto_width(ws_all, num_all)
    ws_all.freeze_panes = "A4"
    ws_all.sheet_view.showGridLines = False

    state_headers_detail = ["Store Name", "Address", "City", "County", "Zip", "Account Type", "Latitude", "Longitude"]
    num_state_cols = len(state_headers_detail)

    TYPE_COLORS_TAB = {
        "Promo Only (Not on C4C)": "DC2626",
        "On Both Lists": "16A34A",
        "C4C Only": "2563EB",
    }

    used_tabs = set()
    for sc in sorted(by_state.keys()):
        accts = by_state[sc]
        state_name = US_STATE_NAMES.get(sc, sc)
        if sc in US_STATE_NAMES or len(sc) <= 3:
            tab_name = sc
        else:
            tab_name = sc[:28]
        if len(tab_name) > 31:
            tab_name = tab_name[:31]
        if tab_name in used_tabs:
            tab_name = f"{tab_name[:28]}_{len(used_tabs)}"
        used_tabs.add(tab_name)

        ws = wb.create_sheet(tab_name)

        primary_type = max(
            ["Promo Only (Not on C4C)", "On Both Lists", "C4C Only"],
            key=lambda t: sum(1 for a in accts if a.get("type") == t)
        )
        ws.sheet_properties.tabColor = TYPE_COLORS_TAB.get(primary_type, "64748B")

        ws.merge_cells(f"A1:{get_column_letter(num_state_cols)}1")
        ws["A1"] = f"{state_name} ({sc}) — {len(accts)} Accounts"
        ws["A1"].font = TITLE_FONT
        ws.row_dimensions[1].height = 32

        promo_ct = sum(1 for a in accts if a.get("type") == "Promo Only (Not on C4C)")
        both_ct = sum(1 for a in accts if a.get("type") == "On Both Lists")
        c4c_ct = sum(1 for a in accts if a.get("type") == "C4C Only")
        counties_in_state = len(set(a.get("county", "") for a in accts if a.get("county")))

        ws.cell(row=2, column=1, value="Counties:").font = BOLD_FONT
        ws.cell(row=2, column=2, value=counties_in_state).font = DATA_FONT
        ws.cell(row=2, column=3, value="Promo Only:").font = BOLD_FONT
        ws.cell(row=2, column=3).fill = RED_FILL
        ws.cell(row=2, column=4, value=promo_ct).font = DATA_FONT
        ws.cell(row=2, column=5, value="Both Lists:").font = BOLD_FONT
        ws.cell(row=2, column=5).fill = GREEN_FILL
        ws.cell(row=2, column=6, value=both_ct).font = DATA_FONT
        ws.cell(row=2, column=7, value="C4C Only:").font = BOLD_FONT
        ws.cell(row=2, column=7).fill = BLUE_FILL
        ws.cell(row=2, column=8, value=c4c_ct).font = DATA_FONT

        row = 4
        for ci, h in enumerate(state_headers_detail, 1):
            ws.cell(row=row, column=ci, value=h)
        _apply_header_row(ws, row, num_state_cols)
        header_row_s = row
        row += 1

        for a in sorted(accts, key=lambda x: (x.get("county", ""), x.get("store_name", ""))):
            ws.cell(row=row, column=1, value=a.get("store_name", ""))
            ws.cell(row=row, column=2, value=a.get("address", ""))
            ws.cell(row=row, column=3, value=a.get("city", ""))
            ws.cell(row=row, column=4, value=a.get("county", ""))
            ws.cell(row=row, column=5, value=a.get("zip", ""))
            acct_type = a.get("type", "")
            ws.cell(row=row, column=6, value=acct_type)
            ws.cell(row=row, column=7, value=a.get("latitude", ""))
            ws.cell(row=row, column=8, value=a.get("longitude", ""))

            _apply_data_row(ws, row, num_state_cols, alt=(row % 2 == 0))

            tf = _type_fill(acct_type)
            if tf:
                ws.cell(row=row, column=6).fill = tf

            row += 1

        _add_autofilter(ws, header_row_s, num_state_cols, row - 1)
        _auto_width(ws, num_state_cols)
        ws.freeze_panes = "A5"
        ws.sheet_view.showGridLines = False

    county_ws = wb.create_sheet("County Summary")
    county_ws.sheet_properties.tabColor = GOLD

    county_ws.merge_cells("A1:G1")
    county_ws["A1"] = f"County-Level Summary ({len(county_details)} counties)"
    county_ws["A1"].font = TITLE_FONT
    county_ws.row_dimensions[1].height = 32

    row = 3
    c_headers = ["County", "State", "Total", "Promo Only", "On Both Lists", "C4C Only", "Gap %"]
    for ci, h in enumerate(c_headers, 1):
        county_ws.cell(row=row, column=ci, value=h)
    _apply_header_row(county_ws, row, len(c_headers))
    header_row_county = row
    row += 1

    for (county, state), stats in sorted(county_details.items(), key=lambda x: (-x[1]["total"], x[0][1], x[0][0])):
        gap = stats["promo"] / stats["total"] * 100 if stats["total"] > 0 else 0
        county_ws.cell(row=row, column=1, value=county)
        county_ws.cell(row=row, column=2, value=state)
        county_ws.cell(row=row, column=3, value=stats["total"])
        county_ws.cell(row=row, column=4, value=stats["promo"])
        county_ws.cell(row=row, column=5, value=stats["both"])
        county_ws.cell(row=row, column=6, value=stats["c4c"])
        county_ws.cell(row=row, column=7, value=f"{gap:.1f}%")

        _apply_data_row(county_ws, row, len(c_headers), alt=(row % 2 == 0))
        for col in [2, 3, 4, 5, 6, 7]:
            county_ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

        if gap >= 80:
            county_ws.cell(row=row, column=7).fill = RED_FILL
            county_ws.cell(row=row, column=7).font = Font(name="Calibri", bold=True, size=10, color=RED_ACCENT)
        elif gap >= 50:
            county_ws.cell(row=row, column=7).fill = AMBER_FILL

        row += 1

    _add_autofilter(county_ws, header_row_county, len(c_headers), row - 1)
    _auto_width(county_ws, len(c_headers))
    county_ws.freeze_panes = "A4"
    county_ws.sheet_view.showGridLines = False

    dist_list = [c for c in customers if c.get("type") == "Distributor"]
    if dist_list:
        ws_dist = wb.create_sheet("Distributors")
        ws_dist.sheet_properties.tabColor = "F59E0B"

        dist_headers = ["Name", "Address", "City", "State", "County", "Zip"]
        num_dh = len(dist_headers)

        ws_dist.merge_cells(f"A1:{get_column_letter(num_dh)}1")
        ws_dist["A1"] = f"Butler Performance Distributors ({len(dist_list)})"
        ws_dist["A1"].font = TITLE_FONT
        ws_dist.row_dimensions[1].height = 32

        ws_dist.merge_cells(f"A2:{get_column_letter(num_dh)}2")
        ws_dist["A2"] = f"Distributor locations across {len(set(d.get('state','') for d in dist_list))} states"
        ws_dist["A2"].font = Font(name="Calibri", size=11, color="64748B")

        row = 4
        for ci, h in enumerate(dist_headers, 1):
            ws_dist.cell(row=row, column=ci, value=h)
        _apply_header_row(ws_dist, row, num_dh)
        header_row_dist = row
        row += 1

        dist_by_state = {}
        for d in sorted(dist_list, key=lambda x: (x.get("state", ""), x.get("store_name", ""))):
            st = d.get("state", "")
            if st not in dist_by_state:
                dist_by_state[st] = []
            dist_by_state[st].append(d)

        for st in sorted(dist_by_state.keys()):
            state_name = US_STATE_NAMES.get(st, st)
            ws_dist.merge_cells(f"A{row}:{get_column_letter(num_dh)}{row}")
            ws_dist.cell(row=row, column=1, value=f"{state_name} ({len(dist_by_state[st])})")
            ws_dist.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=10, color=PURPLE)
            ws_dist.cell(row=row, column=1).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            for col in range(1, num_dh + 1):
                ws_dist.cell(row=row, column=col).border = THIN_BORDER
            row += 1

            for d in dist_by_state[st]:
                ws_dist.cell(row=row, column=1, value=d.get("store_name", ""))
                ws_dist.cell(row=row, column=2, value=d.get("address", ""))
                ws_dist.cell(row=row, column=3, value=d.get("city", ""))
                ws_dist.cell(row=row, column=4, value=d.get("state", ""))
                ws_dist.cell(row=row, column=5, value=d.get("county", ""))
                ws_dist.cell(row=row, column=6, value=d.get("zip", ""))

                _apply_data_row(ws_dist, row, num_dh, alt=(row % 2 == 0))
                ws_dist.cell(row=row, column=4).alignment = Alignment(horizontal="center")
                row += 1

        _add_autofilter(ws_dist, header_row_dist, num_dh, row - 1)
        _auto_width(ws_dist, num_dh)
        ws_dist.column_dimensions["A"].width = 40
        ws_dist.column_dimensions["B"].width = 35
        ws_dist.freeze_panes = "A5"
        ws_dist.sheet_view.showGridLines = False

    wb.save(output_path)

    return {
        "total": len(customers),
        "states": len(by_state),
        "counties": len(county_details),
        "sheets": len(wb.sheetnames),
        "installers": installer_count,
        "distributors": distributor_count,
        "powersports": ps_count,
        "international": intl_count,
        "canada": ca_count,
    }
